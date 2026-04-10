"""Parser tests for data.py.

Targets the three most likely client-facing failure modes:
  1. Header-row offset drift — extra banner row shifts the data block.
  2. Column-shift breakage — inserting a column mis-maps brand/usage.
  3. Sheet-name mismatch — trimmed equality fails on stray spaces.

Builds synthetic minimal xlsx bytes via openpyxl; no bundled fixtures.
"""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

import io
import pandas as pd
import pytest
from openpyxl import Workbook


def _write_xlsx(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _water_rows(header_rows: int = 5, extra_col_before_brand: int = 0):
    """Build a canonical 수도 사용 내역 workbook.

    header_rows controls how many rows precede the data block (the
    parser expects exactly 5: row 0 section title + rows 2-4 headers).
    Insert extra_col_before_brand blank columns before the brand
    column (index 9) to simulate column shift.

    Returns: (workbook, sheet_name)
    """
    wb = Workbook()
    ws = wb.active
    sheet_name = "수도 사용 내역"
    ws.title = sheet_name

    # Section title at row 0
    ws.cell(row=1, column=1, value="▣ 수도 사용 내역")

    # Pad out to header_rows total header lines
    for r in range(2, header_rows + 1):
        ws.cell(row=r, column=1, value=f"header_pad_{r}")

    # Data rows start at header_rows + 1
    # Canonical columns (1-indexed in openpyxl):
    #   col 3 (idx 2)  = building
    #   col 4 (idx 3)  = floor
    #   col 5 (idx 4)  = unit
    #   col 6 (idx 5)  = size_m2
    #   col 9 (idx 8)  = size_py
    #   col 10 (idx 9) = brand      <- shift inserts blanks before this
    #   col 11 (idx 10) = usage_m3
    #   col 21 (idx 20) = total
    brand_col = 10 + extra_col_before_brand  # 1-indexed
    usage_col = 11 + extra_col_before_brand
    total_col = 21 + extra_col_before_brand

    data = [
        ("A", "1F", "101", 50.0, 15.1, "깨비옥",     12.5, 45000),
        ("A", "2F", "201", 60.0, 18.1, "올리브영",   20.0, 60000),
        ("B", "1F", "101", 45.0, 13.6, "스타벅스",    8.0, 30000),
    ]
    for i, (bldg, fl, unit, m2, py, brand, usage, total) in enumerate(data):
        r = header_rows + 1 + i
        ws.cell(row=r, column=3, value=bldg)
        ws.cell(row=r, column=4, value=fl)
        ws.cell(row=r, column=5, value=unit)
        ws.cell(row=r, column=6, value=m2)
        ws.cell(row=r, column=9, value=py)
        ws.cell(row=r, column=brand_col, value=brand)
        ws.cell(row=r, column=usage_col, value=usage)
        ws.cell(row=r, column=total_col, value=total)

    return wb, sheet_name


# ── Failure mode 1: header-row offset drift ──────────────────────────────────

class TestHeaderOffsetDrift:
    """Parser hardcodes raw.iloc[5:] as the start of the data block.

    The parser is resilient in ONE direction (extra rows above data get
    filtered out by the building filter) but fragile in the OTHER
    direction (missing header row causes the first data row to be
    skipped silently). Document both.
    """

    def test_canonical_layout_parses_three_brands(self):
        """Sanity: canonical layout returns the three planted brands."""
        from data import read_water_sheet
        wb, sheet = _water_rows(header_rows=5)
        df = read_water_sheet("test.xlsx", _write_xlsx(wb), sheet)
        assert len(df) == 3
        assert set(df["brand"]) == {"깨비옥", "올리브영", "스타벅스"}

    def test_extra_banner_row_is_absorbed(self):
        """Resilient: one extra banner row above headers is filtered out.

        iloc[5:] is open-ended. The extra row at position 5 gets rejected
        by the building filter (NaN in col 2), and data rows 6-8 survive.
        This documents that the parser tolerates one extra banner row.
        """
        from data import read_water_sheet
        wb, sheet = _water_rows(header_rows=6)
        df = read_water_sheet("test.xlsx", _write_xlsx(wb), sheet)
        assert len(df) == 3

    def test_missing_header_row_drops_first_brand(self):
        """Fragile: one fewer header row causes the first data row to be lost.

        With header_rows=4, the first data row sits at pandas index 4
        (0-indexed). iloc[5:] skips it silently. Only 2 of 3 rows
        survive — a real bug for client files where the section title
        is merged or a header row is missing.
        """
        from data import read_water_sheet
        wb, sheet = _water_rows(header_rows=4)
        df = read_water_sheet("test.xlsx", _write_xlsx(wb), sheet)
        # The first planted brand (깨비옥) is silently dropped.
        assert "깨비옥" not in set(df["brand"].astype(str))
        assert len(df) == 2


# ── Failure mode 2: column-shift breakage ────────────────────────────────────

class TestColumnShift:
    """All parsers use fixed column indices — they ignore header labels.

    Inserting a blank column before the brand column silently maps the
    wrong cell into the brand field. This is the most dangerous failure
    because the client never sees an error — they see wrong data.
    """

    def test_blank_column_before_brand_misreads_brand(self):
        """Shifting brand right by one column should cause mis-map."""
        from data import read_water_sheet
        wb, sheet = _water_rows(header_rows=5, extra_col_before_brand=1)
        df = read_water_sheet("test.xlsx", _write_xlsx(wb), sheet)
        # With the shift, col 9 (where parser looks for brand) is now
        # blank, so no rows should have the planted brand names.
        if len(df) > 0:
            assert "깨비옥" not in set(df["brand"].astype(str))
            assert "올리브영" not in set(df["brand"].astype(str))


# ── Failure mode 3: sheet-name mismatch ──────────────────────────────────────

class TestSheetNameMismatch:
    """`_find_sheet` in app.py uses trimmed equality.

    Real client files sometimes have the sheet named 수도 사용내역
    (no space between 사용 and 내역), or have an autosaved suffix like
    수도 사용 내역(1). Both fail the exact match and the parser is
    never called — the sheet is silently skipped.
    """

    def test_exact_match(self):
        from app import _find_sheet
        sheets = ["검침 내역", "수도 사용 내역", "온수 사용 내역"]
        assert _find_sheet(sheets, "수도 사용 내역") == "수도 사용 내역"

    def test_missing_internal_space_fails(self):
        """수도 사용내역 (no space) does not match 수도 사용 내역."""
        from app import _find_sheet
        sheets = ["수도 사용내역"]
        assert _find_sheet(sheets, "수도 사용 내역") is None

    def test_trailing_parenthesis_suffix_fails(self):
        """수도 사용 내역(1) (autosave suffix) does not match."""
        from app import _find_sheet
        sheets = ["수도 사용 내역(1)"]
        assert _find_sheet(sheets, "수도 사용 내역") is None

    def test_leading_trailing_whitespace_is_trimmed(self):
        """The trimmed equality should succeed on pure whitespace padding."""
        from app import _find_sheet
        sheets = ["  수도 사용 내역  "]
        assert _find_sheet(sheets, "수도 사용 내역") == "  수도 사용 내역  "
